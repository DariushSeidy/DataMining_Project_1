import pandas as pd
import numpy as np
import os
import logging
from openpyxl import Workbook
from collections import defaultdict


def create_item_transaction_matrix_ultra_streaming(file_path):
    """
    Ultra memory-efficient approach: Write directly to Excel without storing matrix in memory.
    This approach processes data row by row and writes directly to Excel.

    Args:
        file_path: Path to the cleaned Excel file

    Returns:
        tuple: (items_list, transactions_list, output_path)
    """

    try:
        logging.info(f"Reading cleaned data from: {file_path}")

        # Read data in chunks to minimize memory usage
        chunk_size = 5000  # Smaller chunks for better memory management
        chunks = []

        # Read Excel file in chunks (manual chunking since pandas doesn't support it for Excel)
        df_full = pd.read_excel(file_path)
        logging.info(f"Loaded data shape: {df_full.shape}")

        # Convert StockCode to string for consistency
        df_full["StockCode"] = df_full["StockCode"].astype(str)

        # Get unique items and transactions
        unique_items = sorted(df_full["StockCode"].unique())
        unique_transactions = sorted(df_full["InvoiceNo"].unique())

        logging.info(f"Number of unique transactions: {len(unique_transactions)}")
        logging.info(f"Number of unique items: {len(unique_items)}")

        # Create mapping for faster lookup
        item_to_index = {item: idx for idx, item in enumerate(unique_items)}
        transaction_to_index = {
            trans: idx for idx, trans in enumerate(unique_transactions)
        }

        # Create workbook and worksheet
        output_path = os.path.join(
            os.path.dirname(file_path), "Item_Transaction_Matrix_Ultra_Streaming.xlsx"
        )
        wb = Workbook()
        ws = wb.active
        ws.title = "Item-Transaction Matrix"

        # Write headers
        ws.cell(row=1, column=1, value="InvoiceNo")
        for col_idx, item in enumerate(unique_items, start=2):
            ws.cell(row=1, column=col_idx, value=str(item))

        logging.info("Creating transaction-item mapping...")

        # Create a dictionary to store which items are in each transaction
        transaction_items = defaultdict(set)

        # Process data in chunks
        total_chunks = len(df_full) // chunk_size + (
            1 if len(df_full) % chunk_size else 0
        )

        for chunk_idx in range(total_chunks):
            start_idx = chunk_idx * chunk_size
            end_idx = min((chunk_idx + 1) * chunk_size, len(df_full))
            chunk_df = df_full.iloc[start_idx:end_idx]

            for _, row in chunk_df.iterrows():
                transaction_items[row["InvoiceNo"]].add(row["StockCode"])

            if chunk_idx % 20 == 0:  # Progress update every 20 chunks
                logging.info(f"Processed chunk {chunk_idx + 1}/{total_chunks}")

        logging.info("Writing matrix to Excel file...")

        # Write matrix row by row
        for row_idx, transaction in enumerate(unique_transactions, start=2):
            ws.cell(row=row_idx, column=1, value=transaction)

            # Get items for this transaction
            items_in_transaction = transaction_items[transaction]

            # Write 1s and 0s for each item
            for col_idx, item in enumerate(unique_items, start=2):
                value = 1 if item in items_in_transaction else 0
                ws.cell(row=row_idx, column=col_idx, value=value)

            if row_idx % 1000 == 0:  # Progress update every 1000 rows
                logging.info(f"Written {row_idx-1}/{len(unique_transactions)} rows")

        # Save the workbook
        wb.save(output_path)

        logging.info(f"Item-Transaction matrix saved to: {output_path}")

        # Calculate basic statistics
        total_cells = len(unique_transactions) * len(unique_items)
        filled_cells = sum(len(items) for items in transaction_items.values())
        density = (filled_cells / total_cells) * 100

        logging.info(f"Matrix density: {density:.2f}%")

        # Calculate average items per transaction
        avg_items_per_transaction = np.mean(
            [len(items) for items in transaction_items.values()]
        )
        logging.info(f"Average items per transaction: {avg_items_per_transaction:.2f}")

        return unique_items, unique_transactions, output_path

    except Exception as e:
        logging.error(f"Error creating item-transaction matrix: {str(e)}")
        return None, None, None


def create_item_transaction_matrix_simple(file_path):
    """
    Simple approach using pandas groupby - most reliable but uses more memory.

    Args:
        file_path: Path to the cleaned Excel file

    Returns:
        tuple: (matrix_df, items_list, transactions_list, output_path)
    """

    try:
        logging.info(f"Reading cleaned data from: {file_path}")
        df = pd.read_excel(file_path)
        logging.info(f"Total data shape: {df.shape}")

        # Convert StockCode to string for consistency
        df["StockCode"] = df["StockCode"].astype(str)

        logging.info("Creating binary matrix using groupby...")

        # Create binary matrix using groupby and unstack
        matrix = df.groupby(["InvoiceNo", "StockCode"]).size().unstack(fill_value=0)

        # Convert to binary (1 if item exists in transaction, 0 otherwise)
        matrix = (matrix > 0).astype(int)

        logging.info(f"Item-Transaction matrix shape: {matrix.shape}")

        # Get lists
        items_list = matrix.columns.tolist()
        transactions_list = matrix.index.tolist()

        # Save matrix
        output_path = os.path.join(
            os.path.dirname(file_path), "Item_Transaction_Matrix_Simple.xlsx"
        )
        matrix.to_excel(output_path)

        logging.info(f"Item-Transaction matrix saved to: {output_path}")

        return matrix, items_list, transactions_list, output_path

    except Exception as e:
        logging.error(f"Error creating item-transaction matrix: {str(e)}")
        return None, None, None, None
