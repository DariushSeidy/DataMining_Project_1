import pandas as pd
import numpy as np
import os
import logging
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows


def create_item_transaction_matrix_streaming(file_path):
    """
    Creates a binary item-transaction matrix by writing directly to Excel file.
    This approach is memory-efficient for large datasets.

    Args:
        file_path: Path to the cleaned Excel file

    Returns:
        tuple: (items_list, transactions_list, output_path)
    """

    try:
        # Read the cleaned data
        logging.info(f"Reading cleaned data from: {file_path}")
        df = pd.read_excel(file_path)

        logging.info(f"Loaded data shape: {df.shape}")

        # Get unique items and transactions
        unique_items = sorted(df["StockCode"].unique())
        unique_transactions = sorted(df["InvoiceNo"].unique())

        logging.info(f"Number of unique transactions: {len(unique_transactions)}")
        logging.info(f"Number of unique items: {len(unique_items)}")

        # Create mapping for faster lookup
        item_to_index = {item: idx for idx, item in enumerate(unique_items)}
        transaction_to_index = {
            trans: idx for idx, trans in enumerate(unique_transactions)
        }

        # Create workbook and worksheet
        output_path = os.path.join(
            os.path.dirname(file_path), "Item_Transaction_Matrix_Streaming.xlsx"
        )
        wb = Workbook()
        ws = wb.active
        ws.title = "Item-Transaction Matrix"

        # Write headers (item codes)
        ws.cell(row=1, column=1, value="InvoiceNo")  # First column for transaction IDs
        for col_idx, item in enumerate(unique_items, start=2):
            ws.cell(row=1, column=col_idx, value=str(item))

        logging.info("Writing matrix data row by row...")

        # Initialize matrix with zeros
        matrix_data = np.zeros((len(unique_transactions), len(unique_items)), dtype=int)

        # Fill matrix by processing data in chunks
        chunk_size = 10000
        total_chunks = len(df) // chunk_size + (1 if len(df) % chunk_size else 0)

        for chunk_idx in range(total_chunks):
            start_idx = chunk_idx * chunk_size
            end_idx = min((chunk_idx + 1) * chunk_size, len(df))
            chunk_df = df.iloc[start_idx:end_idx]

            for _, row in chunk_df.iterrows():
                trans_idx = transaction_to_index[row["InvoiceNo"]]
                item_idx = item_to_index[row["StockCode"]]
                matrix_data[trans_idx, item_idx] = 1

            if chunk_idx % 10 == 0:  # Progress update every 10 chunks
                logging.info(f"Processed chunk {chunk_idx + 1}/{total_chunks}")

        # Write matrix to Excel
        logging.info("Writing matrix to Excel file...")

        # Write transaction IDs and matrix data
        for row_idx, transaction in enumerate(unique_transactions, start=2):
            ws.cell(row=row_idx, column=1, value=transaction)
            for col_idx in range(len(unique_items)):
                ws.cell(
                    row=row_idx,
                    column=col_idx + 2,
                    value=int(matrix_data[row_idx - 2, col_idx]),
                )

            if row_idx % 1000 == 0:  # Progress update every 1000 rows
                logging.info(f"Written {row_idx-1}/{len(unique_transactions)} rows")

        # Save the workbook
        wb.save(output_path)

        logging.info(f"Item-Transaction matrix saved to: {output_path}")

        # Calculate statistics
        total_cells = len(unique_transactions) * len(unique_items)
        filled_cells = np.sum(matrix_data)
        density = (filled_cells / total_cells) * 100

        logging.info(f"Matrix density: {density:.2f}%")
        logging.info(
            f"Average items per transaction: {np.mean(np.sum(matrix_data, axis=1)):.2f}"
        )
        logging.info(
            f"Average transactions per item: {np.mean(np.sum(matrix_data, axis=0)):.2f}"
        )

        return unique_items, unique_transactions, output_path

    except Exception as e:
        logging.error(f"Error creating item-transaction matrix: {str(e)}")
        return None, None, None


def create_item_transaction_matrix_chunked(file_path):
    """
    Alternative approach: Create matrix using pandas but with chunked processing.
    This is a middle ground between memory efficiency and simplicity.

    Args:
        file_path: Path to the cleaned Excel file

    Returns:
        tuple: (matrix_df, items_list, transactions_list, output_path)
    """

    try:
        logging.info(f"Reading cleaned data from: {file_path}")

        # Read data in chunks to reduce memory usage
        chunk_size = 50000
        chunks = []

        for chunk in pd.read_excel(file_path, chunksize=chunk_size):
            chunks.append(chunk)
            logging.info(f"Loaded chunk with {len(chunk)} rows")

        # Combine chunks
        df = pd.concat(chunks, ignore_index=True)
        logging.info(f"Total data shape: {df.shape}")

        # Create binary matrix using pivot (this is still memory intensive but more efficient)
        logging.info("Creating binary matrix...")

        # First, create a binary indicator
        df["item_present"] = 1

        # Pivot to create matrix
        matrix = df.pivot_table(
            index="InvoiceNo",
            columns="StockCode",
            values="item_present",
            fill_value=0,
            aggfunc="max",  # Use max to ensure binary (1 or 0)
        ).astype(int)

        logging.info(f"Item-Transaction matrix shape: {matrix.shape}")

        # Get lists
        items_list = matrix.columns.tolist()
        transactions_list = matrix.index.tolist()

        # Save matrix
        output_path = os.path.join(
            os.path.dirname(file_path), "Item_Transaction_Matrix_Chunked.xlsx"
        )
        matrix.to_excel(output_path)

        logging.info(f"Item-Transaction matrix saved to: {output_path}")

        return matrix, items_list, transactions_list, output_path

    except Exception as e:
        logging.error(f"Error creating item-transaction matrix: {str(e)}")
        return None, None, None, None
